import streamlit as st
import pandas as pd
from fractions import Fraction
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from StyleMap import style_map

# --- Helper: Parse Size ---
def parse_size(size_str):
    if pd.isna(size_str):
        return None
    size_str = str(size_str).strip()
    if ' ' in size_str:
        whole, frac = size_str.split(' ')
        return float(whole) + float(Fraction(frac))
    try:
        return float(size_str)
    except:
        return None

# --- Helper: Process CSV and return Excel file in memory ---
def generate_pivot_excel(uploaded_file):
    df = pd.read_csv(uploaded_file, dtype=str)
    df.columns = df.columns.str.strip()

    # Vendor Style
    df['Vendor Style'] = pd.to_numeric(df['UPC/EAN'].str.strip(), errors='coerce')
    df['KPR Style'] = df['Vendor Style'].map(lambda x: style_map.get(x, {}).get('KPR', 'Unknown Style'))
    df["Mark's Style"] = df['Vendor Style'].map(lambda x: style_map.get(x, {}).get('Marks', 'Unknown Style'))

    # Qty
    df['Qty per Store #'] = pd.to_numeric(df['Qty per Store #'], errors='coerce').fillna(0)
    df['Size'] = df['Size'].apply(parse_size)

    # Extract metadata
    po_number = df['PO Number'].dropna().iloc[0]
    ship_date = df['Ship Dates'].dropna().iloc[0]

    # Pivot
    pivot = pd.pivot_table(
        df,
        index=['Color', 'KPR Style', "Mark's Style"],
        columns=['Size'],
        values='Qty per Store #',
        aggfunc='sum',
        fill_value=0,
        margins=True,
        margins_name='Grand Total'
    ).reset_index()
    pivot = pivot[~pivot['Color'].eq('All')]
    size_cols = sorted([c for c in pivot.columns if c not in ['Color', 'KPR Style', "Mark's Style", 'Grand Total']])
    pivot = pivot[['Color', 'KPR Style', "Mark's Style"] + size_cols + ['Grand Total']]

    # Save to BytesIO instead of file
    output = BytesIO()
    pivot.to_excel(output, index=False)
    output.seek(0)

    # Format with openpyxl
    wb = load_workbook(output)
    ws = wb.active

    # Header row
    ws.insert_rows(1)
    size_start_col = 4
    size_end_col = ws.max_column - 1
    ws.merge_cells(start_row=1, start_column=size_start_col, end_row=1, end_column=size_end_col)
    ws.cell(row=1, column=size_start_col, value="Size").alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=1, column=size_start_col).font = Font(bold=True)
    ws.cell(row=1, column=1, value=f"PO Number: {po_number}").font = Font(bold=True)

    # Auto column width
    for i, col_cells in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value)) for cell in col_cells if cell.value is not None), default=0)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    # Borders + fills
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    accent6_green = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    accent1_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    for cell in ws[1]:
        cell.fill = accent6_green
    for cell in ws[2]:
        cell.fill = accent6_green
        cell.font = Font(bold=True)
    for cell in ws[ws.max_row]:
        cell.fill = accent6_green
        cell.font = Font(bold=True)

    # Alternating row fill
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row-1, min_col=1, max_col=ws.max_column)):
        fill = PatternFill(fill_type=None) if i % 2 == 0 else accent1_blue
        for cell in row:
            cell.fill = fill

    # Ship date row
    ship_date_row = ws.max_row + 1
    ws.cell(row=ship_date_row, column=1, value=f"Ship Date: {ship_date}").font = Font(bold=True)
    ws.cell(row=ship_date_row, column=1).border = border
    ws.cell(row=ship_date_row, column=1).fill = accent6_green

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output, po_number

# --- Streamlit UI ---
st.title("📊 PO Pivot Table Generator")

uploaded_file = st.file_uploader("Upload a PO CSV file", type="csv")

if uploaded_file:
    st.success("File uploaded successfully! Click below to generate.")
    if st.button("Generate Pivot Excel"):
        excel_file, po_number = generate_pivot_excel(uploaded_file)
        st.download_button(
            label="⬇️ Download Pivot Excel",
            data=excel_file,
            file_name=f"PO_{po_number}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
